"""
영양 섭취 또래 비교용 참조 데이터 전처리 스크립트.

질병관리청 "2024 국민건강통계"(국민건강영양조사) 엑셀에서 성별×연령대별
① 영양소 1일 평균 섭취량, ② 영양소 섭취기준 대비비율을 뽑아 작은 JSON으로 만든다.

왜 원시자료(KNHANES raw)가 아니라 국민건강통계(집계표)를 쓰는가?
- 원시자료는 개인 단위 + 복합표본설계라, 표본가중치를 제대로 적용하지 않으면 단순 평균이
  모집단을 대표하지 못한다. 우리가 필요한 건 "30대 남성 평균 단백질 섭취량" 수준의
  집계값뿐이라, 이미 가중치 처리가 끝난 공식 집계표를 쓰는 게 정확하고 간단하다.
- prepare_posture_reference.py가 세종시 집계 데이터를 쓰는 것과 같은 판단이다.

왜 체지방률/골격근량은 없는가?
- 2024 국민건강통계에는 체성분(체지방률·골격근량) 항목이 없다. 신체계측은 신장/체중/
  허리둘레/BMI까지다. 그래서 이 스크립트는 영양 섭취만 다루고, 인바디 해석은 별도 과제다.

엑셀 레이아웃(두 파일 공통):
    8행  : 연도 헤더 ('98 ... '24). 한 연도가 3칸(n / 평균 / (표준오차))을 차지한다.
    10행~: "전체" 블록 → "연령(세)" 소블록 → 거주지역/소득수준 → "남자" 블록 → "여자" 블록
    성별 블록 안의 "연령(세)" 소블록에 우리가 쓸 연령대별 값이 들어 있다.
행 번호를 하드코딩하지 않고 B/C열 라벨을 따라가며 구조적으로 읽는다(시트마다 행이 조금씩 다름).

실행:
    python ml_training/prepare_nutrition_reference.py "<2024국민건강통계 엑셀 폴더 경로>"
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

OUTPUT_PATH = Path(__file__).parent.parent / "app" / "insight" / "data" / "nutrition_reference.json"

SURVEY_YEAR_HEADER = "’24"  # 엑셀이 쓰는 작은따옴표는 일반 '(U+0027)가 아니라 ’(U+2019)다

INTAKE_FILE = "10. 영양소 섭취.xlsx"
RATIO_FILE = "12. 영양소별 섭취기준 대비비율.xlsx"

# 우리 서비스가 실제로 집계하는 4종(DailyTotal의 kcal/protein/carbs/fat)만 가져온다.
# 섭취기준 대비비율은 12번 파일에 에너지·단백질만 있어서 그 둘만 채워진다.
INTAKE_SHEETS = {
    "energy_kcal": "4.에너지",
    "protein_g": "8.단백질",
    "fat_g": "9.지방",
    "carbs_g": "16.탄수화물",
}
RATIO_SHEETS = {
    "energy_ratio_pct": "1.에너지_섭취비율",
    "protein_ratio_pct": "2.단백질_섭취비율",
}

# 성인 사용자를 전제하지만 10대 가입자도 있을 수 있어 10-18까지 포함한다.
AGE_BRACKETS = ["10-18", "19-29", "30-39", "40-49", "50-59", "60-69", "70+"]
SEX_LABELS = {"남자": "M", "여자": "F"}

# " 1-9", " 19-29", " 70+" 처럼 생긴 연령 라벨
AGE_LABEL_PATTERN = re.compile(r"\d+-\d+|\d+\+")


def find_value_column(sheet) -> int:
    """8행의 연도 헤더에서 2024 열을 찾는다. 값(평균/비율)은 그 다음 칸이다."""
    for row in sheet.iter_rows(min_row=8, max_row=8, values_only=True):
        for index, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip() == SURVEY_YEAR_HEADER:
                return index + 1  # n 다음 칸이 평균(또는 비율)
    raise ValueError(f"{SURVEY_YEAR_HEADER} 연도 열을 찾지 못했습니다: {sheet.title}")


def to_float(value):
    """엑셀 셀이 숫자가 아니라 문자열('87.8')로 들어있는 경우가 있어 함께 처리한다.
    값이 억제된 칸('-', '*' 등)은 None으로 흘려보낸다."""
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
            return float(cleaned)
    return None


def read_sheet(sheet) -> dict:
    """{성별코드: {연령대: (값, 표본수)}} 형태로 성별×연령대 값을 읽어온다."""
    value_col = find_value_column(sheet)
    result = {"M": {}, "F": {}}

    current_sex = None
    in_age_block = False
    for row in sheet.iter_rows(min_row=9, values_only=True):
        block_label = str(row[1]).strip() if row[1] else ""
        sub_label = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if block_label in SEX_LABELS:
            current_sex = SEX_LABELS[block_label]
            in_age_block = False
        elif block_label in ("전체",):
            current_sex = None
            in_age_block = False

        if sub_label == "연령(세)":
            in_age_block = True
            continue
        # 연령 블록은 "거주지역"처럼 다른 구분 제목을 만나면 끝난다. 우리가 쓰지 않는
        # 연령대(1-9 등)도 연령 라벨이므로, 블록 종료 판단은 AGE_BRACKETS이 아니라
        # "연령대처럼 생겼는지"로 해야 한다 (안 그러면 1-9에서 블록이 끊긴다)
        if sub_label and in_age_block and not AGE_LABEL_PATTERN.fullmatch(sub_label):
            in_age_block = False

        if not (in_age_block and current_sex and sub_label in AGE_BRACKETS):
            continue

        sample_size = row[value_col - 1] if len(row) > value_col else None
        value = to_float(row[value_col]) if len(row) > value_col else None
        if value is not None:
            result[current_sex][sub_label] = (round(value, 1), to_float(sample_size))

    return result


def main() -> None:
    if len(sys.argv) < 2:
        print(f'사용법: python {Path(__file__).name} "<2024국민건강통계 엑셀 폴더 경로>"')
        raise SystemExit(1)

    base = Path(sys.argv[1])
    groups: dict = {"M": {}, "F": {}}

    for field, sheet_name in INTAKE_SHEETS.items():
        workbook = openpyxl.load_workbook(base / INTAKE_FILE, read_only=True, data_only=True)
        values = read_sheet(workbook[sheet_name])
        workbook.close()
        for sex, per_age in values.items():
            for age, (value, sample_size) in per_age.items():
                entry = groups[sex].setdefault(age, {"sample_size": sample_size})
                entry[field] = value

    for field, sheet_name in RATIO_SHEETS.items():
        workbook = openpyxl.load_workbook(base / RATIO_FILE, read_only=True, data_only=True)
        values = read_sheet(workbook[sheet_name])
        workbook.close()
        for sex, per_age in values.items():
            for age, (value, _) in per_age.items():
                if age in groups[sex]:
                    groups[sex][age][field] = value

    payload = {
        "source": "질병관리청 2024 국민건강통계 (국민건강영양조사)",
        "survey_year": 2024,
        "ratio_baseline": "2020 한국인 영양소 섭취기준(보건복지부) 기준 권장섭취량 대비 비율",
        "groups": groups,
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    total = sum(len(v) for v in groups.values())
    print(f"저장 완료: {OUTPUT_PATH} (성별×연령대 {total}개 그룹)")


if __name__ == "__main__":
    main()
